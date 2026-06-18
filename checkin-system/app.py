from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3
import json
import os
from datetime import datetime
import pandas as pd
import unicodedata
from flask import send_file

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database connection
def get_db():
    db = sqlite3.connect('checkin.db')
    db.row_factory = sqlite3.Row
    db.execute('PRAGMA foreign_keys = ON')
    return db

# Database migration function
def migrate_database():
    """Add new columns if they don't exist"""
    db = get_db()
    cursor = db.cursor()
    
    # Check existing columns
    existing_columns = cursor.execute("PRAGMA table_info(events)").fetchall()
    column_names = [col[1] for col in existing_columns]
    
    # Add is_passed column if it doesn't exist
    if 'is_passed' not in column_names:
        try:
            cursor.execute('ALTER TABLE events ADD COLUMN is_passed BOOLEAN DEFAULT 0')
            print("Added is_passed column to events table")
        except sqlite3.OperationalError as e:
            print(f"Error adding is_passed column: {e}")
    
    # Check if event_columns table exists
    tables = cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='event_columns'").fetchall()
    if not tables:
        # Create event_columns table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS event_columns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER NOT NULL,
                column_name TEXT NOT NULL,
                column_order INTEGER NOT NULL,
                FOREIGN KEY (event_id) REFERENCES events(id) ON DELETE CASCADE
            )
        ''')
        print("Created event_columns table")
    
    # Check if additional_data column exists in guests table
    existing_guest_columns = cursor.execute("PRAGMA table_info(guests)").fetchall()
    guest_column_names = [col[1] for col in existing_guest_columns]
    
    if 'additional_data' not in guest_column_names:
        try:
            cursor.execute('ALTER TABLE guests ADD COLUMN additional_data TEXT')
            print("Added additional_data column to guests table")
        except sqlite3.OperationalError as e:
            print(f"Error adding additional_data column: {e}")
    
    # Migrate old data if needed
    if 'additional_data' in guest_column_names and all(col in guest_column_names for col in ['data1', 'data2', 'data3', 'data4', 'data5']):
        # Check if there's old data to migrate
        old_guests = cursor.execute(
            'SELECT id, data1, data2, data3, data4, data5 FROM guests WHERE additional_data IS NULL'
        ).fetchall()
        
        if old_guests:
            for guest in old_guests:
                additional_data = {}
                if guest['data1']:
                    additional_data['data1'] = guest['data1']
                if guest['data2']:
                    additional_data['data2'] = guest['data2']
                if guest['data3']:
                    additional_data['data3'] = guest['data3']
                if guest['data4']:
                    additional_data['data4'] = guest['data4']
                if guest['data5']:
                    additional_data['data5'] = guest['data5']
                
                if additional_data:
                    cursor.execute(
                        'UPDATE guests SET additional_data = ? WHERE id = ?',
                        (json.dumps(additional_data), guest['id'])
                    )
            print(f"Migrated {len(old_guests)} old guest records")
    
    db.commit()
    db.close()
    print("Database migration completed")

# Initialize database
def init_db():
    with app.app_context():
        db = get_db()
        
        # Create tables if they don't exist
        with app.open_resource('schema.sql', mode='r') as f:
            db.cursor().executescript(f.read())
        
        # Create default users if they don't exist
        admin = db.execute('SELECT * FROM users WHERE username = ?', ('admin',)).fetchone()
        if not admin:
            db.execute(
                'INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                ('admin', generate_password_hash('admin123'), 'admin')
            )
        
        manager = db.execute('SELECT * FROM users WHERE username = ?', ('manager',)).fetchone()
        if not manager:
            db.execute(
                'INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                ('manager', generate_password_hash('manager123'), 'manager')
            )
        
        db.commit()
        db.close()
        
        # Run migration to add new columns
        migrate_database()

# User class for Flask-Login
class User(UserMixin):
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    db.close()
    if user:
        return User(user['id'], user['username'], user['role'])
    return None

# Helper function to process Excel file
def process_excel_file(file_path):
    try:
        # Read Excel file
        df = pd.read_excel(file_path)
        
        # Print available columns for debugging
        print(f"Excel columns found: {list(df.columns)}")
        
        # Create a mapping of possible column names to standard names
        column_mapping = {}
        
        # Define all possible name mappings
        name_variants = ['имя', 'name', 'first_name', 'first name', 'firstname', 'имя гостя']
        surname_variants = ['фамилия', 'surname', 'last_name', 'last name', 'lastname', 'фамилия гостя']
        second_name_variants = ['отчество', 'second_name', 'second name', 'middlename', 'middle_name', 'middle name', 'patronymic']
        
        for col in df.columns:
            col_original = str(col)
            col_lower = col_original.lower().strip()
            
            # Check if this is a name column
            if col_lower in name_variants:
                column_mapping[col] = 'name'
                print(f"Mapped '{col}' -> 'name'")
            # Check if this is a surname column
            elif col_lower in surname_variants:
                column_mapping[col] = 'surname'
                print(f"Mapped '{col}' -> 'surname'")
            # Check if this is a second name column
            elif col_lower in second_name_variants:
                column_mapping[col] = 'second_name'
                print(f"Mapped '{col}' -> 'second_name'")
            else:
                # Keep original column name for additional data
                # Don't lowercase to preserve original formatting
                column_mapping[col] = col_original
        
        # Rename columns based on mapping
        df = df.rename(columns=column_mapping)
        
        # Now check for required columns
        if 'name' not in df.columns and 'surname' not in df.columns:
            raise ValueError(
                f"Excel file must contain 'Фамилия' and 'Имя' columns.\n"
                f"Found columns: {', '.join(df.columns)}\n"
                f"Please ensure your Excel file has columns named 'Фамилия' and 'Имя'"
            )
        elif 'name' not in df.columns:
            raise ValueError(
                f"Excel file must contain an 'Имя' column.\n"
                f"Found columns: {', '.join(df.columns)}\n"
                f"Please add a column named 'Имя'"
            )
        elif 'surname' not in df.columns:
            raise ValueError(
                f"Excel file must contain a 'Фамилия' column.\n"
                f"Found columns: {', '.join(df.columns)}\n"
                f"Please add a column named 'Фамилия'"
            )
        
        # Separate required columns from additional columns
        required_cols = ['name', 'surname']
        if 'second_name' in df.columns:
            required_cols.append('second_name')
        
        # Get additional columns (keep original names from Excel)
        additional_cols = [col for col in df.columns if col not in required_cols]
        
        return df, required_cols, additional_cols
        
    except Exception as e:
        raise Exception(f"Error processing Excel file: {str(e)}")


# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('manager_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        db.close()
        
        if user and check_password_hash(user['password_hash'], password):
            user_obj = User(user['id'], user['username'], user['role'])
            login_user(user_obj)
            flash('Успешно вошли!', 'success')
            
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('manager_dashboard'))
        else:
            flash('Неправильные логин или пароль', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/manager/event/<int:event_id>/export')
@login_required
def export_event_stats(event_id):
    if current_user.role not in ['manager', 'admin']:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    db = get_db()
    
    # Get event details
    event = db.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    if not event:
        flash('Event not found', 'error')
        return redirect(url_for('manager_dashboard'))
    
    # Get column definitions
    try:
        columns = db.execute(
            'SELECT column_name FROM event_columns WHERE event_id = ? ORDER BY column_order',
            (event_id,)
        ).fetchall()
        column_list = [c['column_name'] for c in columns]
    except:
        column_list = ['name', 'surname']
    
    # Get all guests for this event
    guests = db.execute(
        '''SELECT id, name, surname, second_name, additional_data, checked_in, checked_in_at 
           FROM guests 
           WHERE event_id = ? 
           ORDER BY surname, name''',
        (event_id,)
    ).fetchall()
    
    # Get statistics
    total_guests = len(guests)
    checked_in_count = sum(1 for g in guests if g['checked_in'])
    not_checked_in_count = total_guests - checked_in_count
    check_in_percentage = round((checked_in_count / total_guests * 100), 2) if total_guests > 0 else 0
    
    db.close()
    
    # Create Excel file in memory
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # ===== Sheet 1: Summary Statistics =====
    ws_summary = wb.active
    ws_summary.title = "Check-in Summary"
    
    # Title
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = f"Event Check-in Statistics: {event['name']}"
    ws_summary['A1'].font = Font(size=16, bold=True, color='2c3e50')
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # Event info
    ws_summary['A3'] = 'Event Name:'
    ws_summary['B3'] = event['name']
    ws_summary['A3'].font = Font(bold=True)
    
    ws_summary['A4'] = 'Created Date:'
    ws_summary['B4'] = event['created_at']
    ws_summary['A4'].font = Font(bold=True)
    
    ws_summary['A5'] = 'Export Date:'
    ws_summary['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_summary['A5'].font = Font(bold=True)
    
    # Statistics header
    ws_summary['A7'] = 'Statistics'
    ws_summary['A7'].font = Font(size=14, bold=True, color='2c3e50')
    
    # Statistics data
    stats = [
        ('Total Guests', total_guests),
        ('Checked In', checked_in_count),
        ('Not Checked In', not_checked_in_count),
        ('Check-in Percentage', f'{check_in_percentage}%')
    ]
    
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='ffffff')
    
    ws_summary['A9'] = 'Metric'
    ws_summary['B9'] = 'Value'
    ws_summary['A9'].fill = header_fill
    ws_summary['B9'].fill = header_fill
    ws_summary['A9'].font = header_font
    ws_summary['B9'].font = header_font
    
    for idx, (metric, value) in enumerate(stats, start=10):
        ws_summary[f'A{idx}'] = metric
        ws_summary[f'B{idx}'] = value
        if idx % 2 == 0:
            ws_summary[f'A{idx}'].fill = PatternFill(start_color='ecf0f1', end_color='ecf0f1', fill_type='solid')
            ws_summary[f'B{idx}'].fill = PatternFill(start_color='ecf0f1', end_color='ecf0f1', fill_type='solid')
    
    # Adjust column widths for summary
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 20
    
    # ===== Sheet 2: Detailed Guest List =====
    ws_guests = wb.create_sheet("Guest List")
    
    # Headers
    headers = ['#', 'Surname', 'Name']
    if 'second_name' in column_list:
        headers.append('Second Name')
    
    # Add additional columns
    additional_cols = [col for col in column_list if col not in ['name', 'surname', 'second_name']]
    for col in additional_cols:
        headers.append(col.replace('_', ' ').title())
    
    headers.extend(['Check-in Status', 'Check-in Time'])
    
    # Style for headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_guests.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        cell.font = Font(bold=True, color='ffffff', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add data rows
    checked_in_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
    not_checked_fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
    
    for row_idx, guest in enumerate(guests, 2):
        # Row number
        ws_guests.cell(row=row_idx, column=1, value=row_idx - 1)
        
        # Basic info
        ws_guests.cell(row=row_idx, column=2, value=guest['surname'])
        ws_guests.cell(row=row_idx, column=3, value=guest['name'])
        
        col_offset = 4
        if 'second_name' in column_list:
            ws_guests.cell(row=row_idx, column=col_offset, value=guest['second_name'] or '')
            col_offset += 1
        
        # Additional data
        additional_data = {}
        if guest['additional_data']:
            try:
                additional_data = json.loads(guest['additional_data'])
            except:
                pass
        
        for col in additional_cols:
            ws_guests.cell(row=row_idx, column=col_offset, value=additional_data.get(col, ''))
            col_offset += 1
        
        # Check-in status
        status = 'Checked In' if guest['checked_in'] else 'Not Checked In'
        ws_guests.cell(row=row_idx, column=col_offset, value=status)
        
        # Check-in time
        checkin_time = guest['checked_in_at'] if guest['checked_in_at'] else '-'
        ws_guests.cell(row=row_idx, column=col_offset + 1, value=checkin_time)
        
        # Apply row styling based on check-in status
        for col_idx in range(1, len(headers) + 1):
            cell = ws_guests.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            
            if guest['checked_in']:
                cell.fill = checked_in_fill
            else:
                cell.fill = not_checked_fill
    
    # Auto-adjust column widths for guest list
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row in ws_guests.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_value in row:
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 2, 50)
        ws_guests.column_dimensions[column_letter].width = max(adjusted_width, 12)
    
    # Freeze header row
    ws_guests.freeze_panes = 'A2'
    
    # Add autofilter
    ws_guests.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(guests) + 1}"
    
    # ===== Sheet 3: Checked-in Guests Only =====
    ws_checked = wb.create_sheet("Checked-in Guests")
    
    # Copy headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_checked.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
        cell.font = Font(bold=True, color='ffffff', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add only checked-in guests
    checked_guests = [g for g in guests if g['checked_in']]
    for row_idx, guest in enumerate(checked_guests, 2):
        ws_checked.cell(row=row_idx, column=1, value=row_idx - 1)
        ws_checked.cell(row=row_idx, column=2, value=guest['surname'])
        ws_checked.cell(row=row_idx, column=3, value=guest['name'])
        
        col_offset = 4
        if 'second_name' in column_list:
            ws_checked.cell(row=row_idx, column=col_offset, value=guest['second_name'] or '')
            col_offset += 1
        
        additional_data = {}
        if guest['additional_data']:
            try:
                additional_data = json.loads(guest['additional_data'])
            except:
                pass
        
        for col in additional_cols:
            ws_checked.cell(row=row_idx, column=col_offset, value=additional_data.get(col, ''))
            col_offset += 1
        
        ws_checked.cell(row=row_idx, column=col_offset, value='Checked In')
        ws_checked.cell(row=row_idx, column=col_offset + 1, value=guest['checked_in_at'] if guest['checked_in_at'] else '')
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws_checked.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
    
    # Auto-adjust columns for checked-in sheet
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws_checked.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_value in row:
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws_checked.column_dimensions[column_letter].width = max(adjusted_width, 12)
    
    ws_checked.freeze_panes = 'A2'
    ws_checked.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(checked_guests) + 1}"
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"checkin_stats_{event['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# Admin routes
@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    db = get_db()
    
    # Check if is_passed column exists
    columns = db.execute("PRAGMA table_info(events)").fetchall()
    column_names = [col[1] for col in columns]
    
    if 'is_passed' not in column_names:
        events = db.execute(
            'SELECT *, 0 as is_passed FROM events ORDER BY created_at DESC'
        ).fetchall()
    else:
        events = db.execute(
            'SELECT * FROM events ORDER BY created_at DESC'
        ).fetchall()
    
    # Get guest counts for each event
    event_data = []
    for event in events:
        guest_count = db.execute(
            'SELECT COUNT(*) as count FROM guests WHERE event_id = ?',
            (event['id'],)
        ).fetchone()['count']
        
        checked_in_count = db.execute(
            'SELECT COUNT(*) as count FROM guests WHERE event_id = ? AND checked_in = 1',
            (event['id'],)
        ).fetchone()['count']
        
        # Safely get is_passed
        is_passed = event['is_passed'] if 'is_passed' in event.keys() else 0
        
        event_data.append({
            'id': event['id'],
            'name': event['name'],
            'created_at': event['created_at'],
            'is_active': event['is_active'],
            'is_passed': is_passed,
            'guest_count': guest_count,
            'checked_in_count': checked_in_count
        })
    
    db.close()
    return render_template('admin_dashboard.html', events=event_data)

@app.route('/admin/create_event', methods=['GET', 'POST'])
@login_required
def create_event():
    if current_user.role != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        event_name = request.form['event_name']
        excel_file = request.files.get('guest_list')
        
        if not event_name:
            flash('Event name is required', 'error')
            return render_template('create_event.html')
        
        if not excel_file or not excel_file.filename:
            flash('Excel file is required', 'error')
            return render_template('create_event.html')
        
        if not excel_file.filename.endswith(('.xlsx', '.xls')):
            flash('Only Excel files (.xlsx, .xls) are allowed', 'error')
            return render_template('create_event.html')
        
        db = get_db()
        
        try:
            # Save uploaded file
            filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{excel_file.filename}")
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            excel_file.save(file_path)
            
            # Process Excel file
            df, required_cols, additional_cols = process_excel_file(file_path)
            
            cursor = db.cursor()
            
            # Create event
            cursor.execute(
                'INSERT INTO events (name, created_by) VALUES (?, ?)',
                (event_name, current_user.id)
            )
            event_id = cursor.lastrowid
            
            # Save column definitions
            all_columns = required_cols + additional_cols
            for idx, col in enumerate(all_columns):
                cursor.execute(
                    'INSERT INTO event_columns (event_id, column_name, column_order) VALUES (?, ?, ?)',
                    (event_id, col, idx)
                )
            
            # Import guests
            for _, row in df.iterrows():
                # Prepare required data
                name = str(row.get('name', '')).strip()
                surname = str(row.get('surname', '')).strip()
                second_name = str(row.get('second_name', '')).strip() if 'second_name' in df.columns else ''
                
                # Prepare additional data as JSON
                additional_data = {}
                for col in additional_cols:
                    value = row[col]
                    if pd.notna(value):
                        additional_data[col] = str(value).strip()
                    else:
                        additional_data[col] = ''
                
                cursor.execute(
                    '''INSERT INTO guests 
                       (event_id, name, surname, second_name, additional_data) 
                       VALUES (?, ?, ?, ?, ?)''',
                    (event_id, name, surname, second_name, json.dumps(additional_data))
                )
            
            db.commit()
            
            # Clean up uploaded file
            if os.path.exists(file_path):
                os.remove(file_path)
            
            flash(f'Event created successfully with {len(df)} guests imported!', 'success')
            return redirect(url_for('admin_dashboard'))
            
        except Exception as e:
            db.rollback()
            flash(f'Error: {str(e)}', 'error')
            return render_template('create_event.html')
        finally:
            db.close()
    
    return render_template('create_event.html')

@app.route('/admin/edit_event/<int:event_id>', methods=['GET', 'POST'])
@login_required
def edit_event(event_id):
    if current_user.role != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    db = get_db()
    event = db.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    
    if not event:
        flash('Event not found', 'error')
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        event_name = request.form['event_name']
        excel_file = request.files.get('guest_list')
        
        if not event_name:
            flash('Event name is required', 'error')
            return render_template('edit_event.html', event=event)
        
        try:
            # Update event name
            db.execute('UPDATE events SET name = ? WHERE id = ?', (event_name, event_id))
            
            # If new Excel file is uploaded, reset guest data
            if excel_file and excel_file.filename:
                if not excel_file.filename.endswith(('.xlsx', '.xls')):
                    flash('Only Excel files (.xlsx, .xls) are allowed', 'error')
                    return render_template('edit_event.html', event=event)
                
                # Save uploaded file
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{excel_file.filename}")
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file.save(file_path)
                
                # Process Excel file
                df, required_cols, additional_cols = process_excel_file(file_path)
                
                # Delete old guests and columns
                db.execute('DELETE FROM guests WHERE event_id = ?', (event_id,))
                db.execute('DELETE FROM event_columns WHERE event_id = ?', (event_id,))
                
                # Save new column definitions
                all_columns = required_cols + additional_cols
                for idx, col in enumerate(all_columns):
                    db.execute(
                        'INSERT INTO event_columns (event_id, column_name, column_order) VALUES (?, ?, ?)',
                        (event_id, col, idx)
                    )
                
                # Import new guests
                for _, row in df.iterrows():
                    name = str(row.get('name', '')).strip()
                    surname = str(row.get('surname', '')).strip()
                    second_name = str(row.get('second_name', '')).strip() if 'second_name' in df.columns else ''
                    
                    additional_data = {}
                    for col in additional_cols:
                        value = row[col]
                        if pd.notna(value):
                            additional_data[col] = str(value).strip()
                        else:
                            additional_data[col] = ''
                    
                    db.execute(
                        '''INSERT INTO guests 
                           (event_id, name, surname, second_name, additional_data) 
                           VALUES (?, ?, ?, ?, ?)''',
                        (event_id, name, surname, second_name, json.dumps(additional_data))
                    )
                
                # Clean up uploaded file
                if os.path.exists(file_path):
                    os.remove(file_path)
                
                flash(f'Event updated and {len(df)} guests imported!', 'success')
            else:
                flash('Event name updated successfully!', 'success')
            
            db.commit()
            return redirect(url_for('admin_dashboard'))
            
        except Exception as e:
            db.rollback()
            flash(f'Error: {str(e)}', 'error')
            return render_template('edit_event.html', event=event)
        finally:
            db.close()
    
    # Get current columns for display
    try:
        columns = db.execute(
            'SELECT column_name FROM event_columns WHERE event_id = ? ORDER BY column_order',
            (event_id,)
        ).fetchall()
        column_list = [c['column_name'] for c in columns]
    except:
        column_list = []
    
    db.close()
    return render_template('edit_event.html', event=event, columns=column_list)

@app.route('/admin/toggle_event_passed/<int:event_id>', methods=['POST'])
@login_required
def toggle_event_passed(event_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    db = get_db()
    event = db.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    
    if event:
        # Safely get is_passed value
        try:
            current_status = event['is_passed']
        except (KeyError, IndexError):
            current_status = 0
        
        new_status = 0 if current_status else 1
        
        try:
            db.execute('UPDATE events SET is_passed = ? WHERE id = ?', (new_status, event_id))
        except sqlite3.OperationalError:
            # If column doesn't exist, add it first
            db.execute('ALTER TABLE events ADD COLUMN is_passed BOOLEAN DEFAULT 0')
            db.execute('UPDATE events SET is_passed = ? WHERE id = ?', (new_status, event_id))
        
        db.commit()
        
        status_text = 'marked as passed' if new_status else 'marked as active'
        flash(f'Event {status_text} successfully!', 'success')
    else:
        flash('Event not found', 'error')
    
    db.close()
    return redirect(url_for('admin_dashboard'))

# Manager routes
@app.route('/manager/dashboard')
@login_required
def manager_dashboard():
    if current_user.role not in ['manager', 'admin']:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    db = get_db()
    
    # Check if is_passed column exists
    columns = db.execute("PRAGMA table_info(events)").fetchall()
    column_names = [col[1] for col in columns]
    
    if 'is_passed' not in column_names:
        events = db.execute(
            'SELECT *, 0 as is_passed FROM events WHERE is_active = 1 ORDER BY created_at DESC'
        ).fetchall()
    else:
        events = db.execute(
            'SELECT * FROM events WHERE is_active = 1 AND is_passed = 0 ORDER BY created_at DESC'
        ).fetchall()
    
    db.close()
    return render_template('event_select.html', events=events)

@app.route('/manager/event/<int:event_id>')
@login_required
def guest_search(event_id):
    if current_user.role not in ['manager', 'admin']:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    db = get_db()
    event = db.execute('SELECT * FROM events WHERE id = ?', (event_id,)).fetchone()
    
    if not event:
        flash('Event not found', 'error')
        return redirect(url_for('manager_dashboard'))
    
    # Get column definitions
    try:
        columns = db.execute(
            'SELECT column_name FROM event_columns WHERE event_id = ? ORDER BY column_order',
            (event_id,)
        ).fetchall()
        column_list = [c['column_name'] for c in columns]
    except:
        # If event_columns table doesn't exist or no columns defined
        column_list = ['name', 'surname']
    
    # Get all unique first letters of surnames for this event
    letters = db.execute(
        '''SELECT DISTINCT UPPER(SUBSTR(surname, 1, 1)) as letter 
           FROM guests 
           WHERE event_id = ? AND surname != ''
           ORDER BY letter''',
        (event_id,)
    ).fetchall()
    
    db.close()
    return render_template(
        'guest_search.html',
        event=event,
        letters=[l['letter'] for l in letters],
        columns=column_list
    )

@app.route('/api/search_guests/<int:event_id>')
@login_required
def search_guests(event_id):
    search_term = request.args.get('term', '')
    
    # Debug logging
    app.logger.info(f"Search term received: '{search_term}'")
    app.logger.info(f"Search term bytes: {search_term.encode('utf-8')}")
    
    db = get_db()
    
    if search_term and search_term != 'ALL':
        # Use LIKE directly without UPPER() to avoid Unicode issues
        # This works because SQLite LIKE is case-insensitive for ASCII
        # For Unicode, we'll do a case-insensitive comparison in a different way
        
        # Method 1: Direct LIKE comparison (works for both Latin and Cyrillic in most cases)
        search_pattern = f'{search_term}%'
        
        app.logger.info(f"Search pattern: '{search_pattern}'")
        
        # Try the search
        guests = db.execute(
            '''SELECT id, name, surname, second_name, additional_data, checked_in, checked_in_at 
               FROM guests 
               WHERE event_id = ? AND surname LIKE ? 
               ORDER BY surname, name''',
            (event_id, search_pattern)
        ).fetchall()
        
        app.logger.info(f"Found {len(guests)} guests with direct LIKE")
        
        # If no results and search term might be mixed case, try case-insensitive approach
        if len(guests) == 0:
            # Get all guests and filter manually (less efficient but works for Unicode)
            all_guests = db.execute(
                '''SELECT id, name, surname, second_name, additional_data, checked_in, checked_in_at 
                   FROM guests 
                   WHERE event_id = ? 
                   ORDER BY surname, name''',
                (event_id,)
            ).fetchall()
            
            # Manual filtering
            search_upper = search_term.upper()
            guests = [g for g in all_guests if g['surname'].upper().startswith(search_upper)]
            
            app.logger.info(f"Found {len(guests)} guests with manual filtering")
    else:
        # Show all guests
        guests = db.execute(
            '''SELECT id, name, surname, second_name, additional_data, checked_in, checked_in_at 
               FROM guests 
               WHERE event_id = ? 
               ORDER BY surname, name''',
            (event_id,)
        ).fetchall()
    
    # Get column definitions
    try:
        columns = db.execute(
            'SELECT column_name FROM event_columns WHERE event_id = ? ORDER BY column_order',
            (event_id,)
        ).fetchall()
        column_list = [c['column_name'] for c in columns]
    except:
        column_list = ['name', 'surname']
    
    db.close()
    
    # Process guests data
    guest_list = []
    for g in guests:
        additional_data = {}
        if g['additional_data']:
            try:
                additional_data = json.loads(g['additional_data'])
            except:
                additional_data = {}
        
        guest_list.append({
            'id': g['id'],
            'name': g['name'],
            'surname': g['surname'],
            'second_name': g['second_name'] or '',
            'additional_data': additional_data,
            'checked_in': g['checked_in'],
            'checked_in_at': g['checked_in_at'] if g['checked_in_at'] else None
        })
    
    return jsonify({
        'guests': guest_list,
        'columns': column_list
    })

@app.route('/api/checkin/<int:guest_id>', methods=['POST'])
@login_required
def checkin_guest(guest_id):
    if current_user.role not in ['manager', 'admin']:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    data = request.get_json()
    check_in = data.get('check_in', True)
    
    db = get_db()
    try:
        if check_in:
            db.execute(
                'UPDATE guests SET checked_in = 1, checked_in_at = ? WHERE id = ? AND checked_in = 0',
                (datetime.now(), guest_id)
            )
        else:
            db.execute(
                'UPDATE guests SET checked_in = 0, checked_in_at = NULL WHERE id = ? AND checked_in = 1',
                (guest_id,)
            )
        
        db.commit()
        
        if db.total_changes > 0:
            action = 'checked in' if check_in else 'unchecked'
            return jsonify({
                'success': True, 
                'message': f'Guest {action} successfully',
                'checked_in': check_in,
                'checked_in_at': datetime.now().isoformat() if check_in else None
            })
        else:
            status = 'already checked in' if check_in else 'not checked in'
            return jsonify({'success': False, 'message': f'Guest {status}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        db.close()

@app.route('/api/update_guest/<int:guest_id>', methods=['POST'])
@login_required
def update_guest(guest_id):
    if current_user.role not in ['manager', 'admin']:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    data = request.get_json()
    
    db = get_db()
    try:
        # Get current guest data
        guest = db.execute('SELECT * FROM guests WHERE id = ?', (guest_id,)).fetchone()
        
        if not guest:
            return jsonify({'success': False, 'message': 'Guest not found'}), 404
        
        # Update basic fields
        name = data.get('name', guest['name'])
        surname = data.get('surname', guest['surname'])
        second_name = data.get('second_name', guest['second_name'])
        
        # Update additional data
        additional_data = {}
        if guest['additional_data']:
            try:
                additional_data = json.loads(guest['additional_data'])
            except:
                pass
        
        # Update additional fields from the request
        if 'additional_data' in data:
            for key, value in data['additional_data'].items():
                additional_data[key] = value
        
        # Update the database
        db.execute(
            '''UPDATE guests 
               SET name = ?, surname = ?, second_name = ?, additional_data = ? 
               WHERE id = ?''',
            (name, surname, second_name, json.dumps(additional_data), guest_id)
        )
        
        db.commit()
        
        # Fetch updated guest
        updated_guest = db.execute(
            '''SELECT id, name, surname, second_name, additional_data, checked_in, checked_in_at 
               FROM guests WHERE id = ?''',
            (guest_id,)
        ).fetchone()
        
        # Process additional data for response
        updated_additional_data = {}
        if updated_guest['additional_data']:
            try:
                updated_additional_data = json.loads(updated_guest['additional_data'])
            except:
                pass
        
        return jsonify({
            'success': True,
            'message': 'Guest updated successfully',
            'guest': {
                'id': updated_guest['id'],
                'name': updated_guest['name'],
                'surname': updated_guest['surname'],
                'second_name': updated_guest['second_name'] or '',
                'additional_data': updated_additional_data,
                'checked_in': updated_guest['checked_in'],
                'checked_in_at': updated_guest['checked_in_at'] if updated_guest['checked_in_at'] else None
            }
        })
        
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        db.close()



if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)


